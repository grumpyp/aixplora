from typing import BinaryIO
from fastapi import UploadFile

import openpyxl
import csv
import io
import os


def load_xlsx(file: BinaryIO, filename: str, file_meta: UploadFile):

    misc_dir = os.path.join(os.getcwd(), "misc")
    excel_file = io.BytesIO(file.read())
    workbook = openpyxl.load_workbook(excel_file)

    # Open the CSV file for writing
    output_file_path = f"{misc_dir}/{filename}.txt"
    with open(output_file_path, "w") as csvfile:
        writer = csv.writer(csvfile)

        # Iterate over all worksheets
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for row in worksheet.iter_rows(values_only=True):
                # replace newline if possible
                row = [str(cell).strip().replace("\n", "") for cell in row if cell is not None]
                writer.writerow(row)

    return output_file_path, file_meta
